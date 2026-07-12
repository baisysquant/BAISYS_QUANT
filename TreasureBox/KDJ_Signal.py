from __future__ import annotations
import os
import sys
from datetime import datetime as dt
from datetime import timedelta

import pandas as pd
from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from UtilsManager.ConfigParser import Config
from DataCollection.CalendarManager import TradingCalendarAnalyzer
import akshare as ak  # 可选，为整合新闻数据做准备

if __name__ == "__main__":
    # --- 1. 初始化路径
    current_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(current_dir)
    config_path = os.path.join(root_dir, "config.ini")

    try:
        config = Config(config_path)
        print(f"[OK] 成功加载配置文件: {config_path}")
    except Exception as e:
        print(f"[FAIL] 配置加载失败: {e}")
        sys.exit(1)

    # 创建数据库连接 URI
    DB_URI = f"postgresql://{config.DB_USER}:{config.DB_PASSWORD}@{config.DB_HOST}:{config.DB_PORT}/{config.DB_NAME}"

    try:
        engine = create_engine(DB_URI)

        # 测试数据库连接
        with engine.connect() as conn:
            query_test = text("SELECT 1")
            result = conn.execute(query_test)
            fetchone = result.fetchone()
            assert fetchone is not None, "数据库连接失败或查询失败！"
            print("[DB] 数据库连接已激活")
    except Exception as e:
        print(f"[DB ERROR] 数据库连接失败: {e}")
        sys.exit(1)

    # --- 2. 定义报告输出路径
    REPORT_OUTPUT_DIR = os.path.join(os.path.expanduser(config.HOME_DIRECTORY), "cache", "reports")
    os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
    print(f"[DIR] 报告将保存至: {REPORT_OUTPUT_DIR}")

    # --- 3. 定义时间范围
    one_month_ago = (dt.today() - timedelta(days=15)).strftime("%Y-%m-%d")
    print(f"[DATE] 统计周期: {one_month_ago} 至 {dt.now().strftime('%Y-%m-%d')}")

    # --- 4. 查询所有有 KDJ 信号的股票
    query_signal_stocks = text("""
        SELECT DISTINCT stock_code, stock_name
        FROM app_stock_strategy_report
        WHERE kdj_signal IS NOT NULL
        AND archive_date >= :one_month_ago
        ORDER BY stock_code;
    """)
    try:
        df_signal_stocks = pd.read_sql(query_signal_stocks, engine, params={"one_month_ago": one_month_ago})
    except Exception as e:
        print(f"[SQL Error] 错误：无法获取有 KDJ 信号的股票记录: {e}")
        sys.exit(1)

    if df_signal_stocks.empty:
        print("[WARN] 未找到有 KDJ 信号的股票。")
        sys.exit(1)

    # --- 5. 添加交易所前缀（sh/ sz）
    stock_info_map = {}
    for _, row in df_signal_stocks.iterrows():
        code = str(row["stock_code"]).strip()
        name = row["stock_name"] if row["stock_name"] else "未知名称"
        if code.startswith("6"):
            symbol = f"sh{code}"
        elif code.startswith("0") or code.startswith("3") or code.startswith("8"):
            symbol = f"sz{code}"
        elif code.startswith("8") or code.startswith("4") or code.startswith("9"):
            symbol = f"bj{code}"
        else:
            symbol = f"sz{code}"  # 默认归为深市股票
        stock_info_map[code] = {"name": name, "symbol": symbol}

    # --- 6. 查询有效股票（有价格数据）
    signal_stock_codes = list(stock_info_map.keys())
    stock_symbols = [stock_info_map[code]["symbol"] for code in signal_stock_codes]

    query_valid_symbols = text("""
        SELECT symbol
        FROM stock_daily_kline
        WHERE symbol IN :symbols
        AND trade_date >= :one_month_ago
        AND "close" IS NOT NULL;
    """)
    try:
        df_valid_symbols = pd.read_sql(query_valid_symbols, engine, params={"symbols": tuple(stock_symbols), "one_month_ago": one_month_ago})
    except Exception as e:
        print(f"[SQL Error] 错误：无法获取有价格数据的股票记录: {e}")
        sys.exit(1)

    valid_symbols_set = set(df_valid_symbols["symbol"].tolist())
    effective_stock_codes = [code for code in signal_stock_codes if stock_info_map[code]["symbol"] in valid_symbols_set]

    if not effective_stock_codes:
        print("[WARN] 没有股票同时满足有 KDJ 信号和有价格数据的条件。")
        sys.exit(1)

    # --- 7. 查询信号日收盘价
    query_last_signal_dates = text("""
        SELECT stock_code, MAX(archive_date)::date AS last_signal_date
        FROM app_stock_strategy_report
        WHERE kdj_signal IS NOT NULL
        AND archive_date >= :one_month_ago
        GROUP BY stock_code;
    """)
    try:
        df_last_signal_dates = pd.read_sql(query_last_signal_dates, engine, params={"one_month_ago": one_month_ago})
    except Exception as e:
        print(f"[SQL Error] 错误：无法获取 KDJ 信号日: {e}")
        sys.exit(1)

    # 构建 stock_code 到 signal_date 的映射
    signal_date_map = {
        row["stock_code"]: row["last_signal_date"].strftime("%Y-%m-%d")
        for _, row in df_last_signal_dates.iterrows()
    }

    # 查询信号日收盘价
    query_signal_close = text("""
        SELECT symbol, trade_date, "close"
        FROM stock_daily_kline
        WHERE (symbol, trade_date) IN :signal_date_pairs
        AND "close" IS NOT NULL;
    """)
    signal_date_pairs = [(stock_info_map[code]["symbol"], signal_date_map[code]) for code in effective_stock_codes]
    in_clause_dates = tuple(signal_date_pairs)

    try:
        df_signal_close = pd.read_sql(query_signal_close, engine, params={"signal_date_pairs": in_clause_dates})
    except Exception as e:
        print(f"[SQL Error] 错误：无法获取 KDJ 信号日的收盘价: {e}")
        sys.exit(1)

    # 构建 stock_code 到 signal_close 的映射
    signal_close_map = {}
    for _, row in df_signal_close.iterrows():
        symbol = row["symbol"]
        try:
            date_str = row["trade_date"].strftime("%Y-%m-%d")  # 确保 trade_date 是 datetime 类型
        except Exception:
            date_str = row["trade_date"]
        close_val = row["close"]
        signal_close_map[symbol] = close_val

    # --- 8. 查询最新收盘价（近30天）
    query_latest_close = text("""
        WITH latest_trades AS (
            SELECT symbol, MAX(trade_date) AS latest_date
            FROM stock_daily_kline
            WHERE symbol IN :symbols
            AND trade_date >= :one_month_ago
            AND "close" IS NOT NULL
            GROUP BY symbol
        )
        SELECT lt.symbol, lt.latest_date, sdk."close" AS latest_close
        FROM latest_trades lt
        JOIN stock_daily_kline sdk ON lt.symbol = sdk.symbol AND lt.latest_date = sdk.trade_date;
    """)
    try:
        df_latest_close = pd.read_sql(query_latest_close, engine, params={"symbols": tuple(stock_symbols), "one_month_ago": one_month_ago})
    except Exception as e:
        print(f"[SQL Error] 错误：无法获取最新收盘价: {e}")
        sys.exit(1)

    # 构建 latest_close_map
    latest_close_map = {row["symbol"]: row["latest_close"] for _, row in df_latest_close.iterrows()}

    # --- 9. 最终筛选：信号日后股价是否上涨
    final_effective_stock_codes = []
    filtered_out = []

    for code in effective_stock_codes:
        symbol = stock_info_map[code]["symbol"]
        if code not in signal_close_map or symbol not in latest_close_map:
            filtered_out.append(f"{stock_info_map[code]['name']} ({symbol}) - 无信号日价格或无最新价格。")
            continue

        signal_close = signal_close_map[code]
        latest_close = latest_close_map[symbol]
        gain_pct = ((latest_close - signal_close) / signal_close) * 100
        gain_percentage_map = {code: round(gain_pct, 2) for code in effective_stock_codes}

        if latest_close > signal_close:
            final_effective_stock_codes.append(code)
        else:
            filtered_out.append(
                f"{stock_info_map[code]['name']} ({symbol}) - 信号日 {signal_close}, 最新价 {latest_close} [FAIL] 未上涨"
            )

    if not final_effective_stock_codes:
        print("[WARN] 没有任何股票满足价格上涨的条件。")
        sys.exit(1)

    print(f"[OK] 最终通过趋势验证的有效股票：{len(final_effective_stock_codes)} 只")
    for reason in filtered_out:
        print(f" - {reason}")

    # --- 10. 获取近30天的交易数据（包含所有有效股票）
    query_kline = text("""
        SELECT symbol, trade_date, "close"
        FROM stock_daily_kline
        WHERE symbol IN :symbols
        AND trade_date >= :one_month_ago
        AND "close" IS NOT NULL
        ORDER BY symbol, trade_date;
    """)
    # 注意：这里确保我们传入的是 tuple
    try:
        df_kline = pd.read_sql(query_kline, engine, params={"symbols": tuple(stock_symbols), "one_month_ago": one_month_ago})
    except Exception as e:
        print(f"[SQL Error] 错误：获取交易数据失败: {e}")
        sys.exit(1)

    if df_kline.empty:
        print("[WARN] 未找到有效股票的交易数据（近30天）。")
        sys.exit(1)

    # 转换为 datetime 对象（避免后续 string 使用错误）
    df_kline["trade_date"] = pd.to_datetime(df_kline["trade_date"])

    # 获取并格式化的交易日期
    trade_date_list = [date.strftime("%Y-%m-%d") for date in df_kline["trade_date"].unique()]

    # 打印调试信息
    print(f"[OK] 共获取 {len(trade_date_list)} 个交易日，覆盖范围：{trade_date_list[0]} 至 {trade_date_list[-1]}")

    # 构建 close_map：stock_code 到每一天的收盘价
    close_map = {}
    for _, row in df_kline.iterrows():
        symbol = row["symbol"]
        date_str = row["trade_date"].strftime("%Y-%m-%d")
        close_val = row["close"]
        close_map.setdefault(symbol, {})[date_str] = close_val

    # --- 11. 获取 MACD 信号点
    query_macd_signals = text("""
        SELECT stock_code, archive_date::date AS archive_date, macd_12269_signal
        FROM app_stock_strategy_report
        WHERE macd_12269_signal IS NOT NULL
        AND archive_date >= :one_month_ago
        ORDER BY stock_code, archive_date;
    """)
    try:
        df_macd_signals = pd.read_sql(query_macd_signals, engine, params={"one_month_ago": one_month_ago})
    except Exception as e:
        print(f"[SQL Error] 错误：获取 MACD 信号失败: {e}")
        sys.exit(1)

    # 构建 MACD 高亮映射
    macd_highlight_map = {}
    for _, row in df_macd_signals.iterrows():
        stock_code = str(row["stock_code"]).strip()
        date_str = row["archive_date"].strftime("%Y-%m-%d")
        signal_value = str(row["macd_12269_signal"]).strip()
        symbol = stock_info_map[stock_code]["symbol"]

        if symbol in close_map and date_str in close_map[symbol]:
            if signal_value in ["下金叉", "下叉", "金叉", "buy", "1", "BUY", "正金叉"]:
                color_hex = "ADD8E6"  # 蓝色
            elif signal_value in ["上金叉", "上叉", "死叉", "sell", "-1", "SELL", "负金叉"]:
                color_hex = "9370DB"  # 紫色
            else:
                color_hex = "FF0000"  # 红色
            macd_highlight_map[(symbol, date_str)] = color_hex

    print(f"[OK] 共有 {len(macd_highlight_map)} 个 MACD 信号点可用于高亮")

    # --- 12. 生成 Excel 报告
    print("[REPORT] 正在生成 Excel 报告...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Close & Signal Report"

    # --- 标题行合并
    ws.insert_rows(1)
    title_cell = ws.cell(
        row=1, column=1, value="股票收盘价与多因子信号聚焦报告（近30天，仅展示有信号且价格持续上涨的股票）"
    )
    title_cell.font = Font(bold=True, size=16, color="2E5488")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(trade_date_list) + 3)
    ws.row_dimensions[1].height = 35

    # --- 图例说明行合并
    ws.insert_rows(2)
    note_cell = ws.cell(
        row=2, column=1,
        value="[-] 筛选逻辑：仅展示‘有 KDJ 信号’且‘信号后股价上涨’的股票。\n"
              "[-] 高亮说明：\n"
              "[蓝色] MACD 零轴下金叉（买入信号）\n"
              "[紫色] MACD 零轴上金叉（买入信号）\n"
              "[红色] KDJ 信号（买入/卖出）\n"
              "[OK] 所有股票均满足：信号日后价格上涨，确保动能有效。"
    )
    note_cell.font = Font(bold=True, color="2E5488", size=12)
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=len(trade_date_list) + 3)
    ws.row_dimensions[2].height = 80

    # --- 表头定义
    ws.cell(row=5, column=1, value="Stock Code")
    ws.cell(row=5, column=2, value="Stock Name")
    ws.cell(row=5, column=3, value="Signal to Latest Gain (%)")

    for col_idx, date_str in enumerate(trade_date_list, 4):
        ws.cell(row=5, column=col_idx, value=date_str)

    # --- 填充数据行
    row_idx = 6
    for code in final_effective_stock_codes:
        symbol = stock_info_map[code]["symbol"]
        name = stock_info_map[code]["name"]
        gain_pct = gain_percentage_map[code]

        # 写入股票代码、股票名称、涨幅
        ws.cell(row=row_idx, column=1, value=code)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=f"{gain_pct:+.2f}%")

        # 写入每个交易日的收盘价，设置颜色高亮
        for col_idx, date_str in enumerate(trade_date_list, 4):
            if date_str in close_map[symbol]:
                cell = ws.cell(row=row_idx, column=col_idx, value=close_map[symbol][date_str])
                if (symbol, date_str) in macd_highlight_map:
                    color_hex = macd_highlight_map[(symbol, date_str)]
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    cell.fill = fill
            else:
                ws.cell(row=row_idx, column=col_idx, value=None)

        row_idx += 1

    # --- 自动调整列宽
    if trade_date_list:
        total_cols = len(trade_date_list) + 3
        for col_idx in range(1, total_cols + 1):
            column = get_column_letter(col_idx)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max(max_length + 2, 8), 25)

    # --- 保存 Excel 文件
    calendar_mgr = TradingCalendarAnalyzer()
    today_str = calendar_mgr.get_last_trading_day()
    excel_file = os.path.join(REPORT_OUTPUT_DIR, f"KDJ报告_{today_str}.xlsx")

    try:
        wb.save(excel_file)
        print(f"[DONE] Excel 文件已生成：{excel_file}")
    except Exception as e:
        print(f"[ERROR] Excel 保存失败: {e}")
        sys.exit(1)

    # 最后打印变量帮助调试
    print(f"[INFO] trade_date_list: {trade_date_list}")
    print(f"[INFO] one_month_ago: {one_month_ago}")